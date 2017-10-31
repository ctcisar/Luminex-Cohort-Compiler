import openpyxl
import configparser
import os
from numpy import average
from numpy import std
from numpy import transpose

# Config vars

config = configparser.ConfigParser()
config.read('scriptconfig.ini')

# mandatory input vars

PROTOCOL_NAME = config['input']['name']
PLATE_COUNT = int(config['input']['plate_count'])
PLATE_TYPES = config['input']['plate_types'].split(" ")

def check_and_default(config,cat,key,default):
    if key in config[cat].keys():
        return config[cat][key]
    else:
        return default

def is_number(s):
    if s is None:
        return False
    try:
        float(s)
        return True
    except ValueError:
        return False

# analysis vars

BEAD_CUTOFF = int(check_and_default(config,"analysis","bead_cutoff","25"))
CONTROL_NAMES = check_and_default(config,"analysis","control_names",False)
if CONTROL_NAMES is not False:
    CONTROL_NAMES = CONTROL_NAMES.split(",")
COMBINE_CONTROLS = (check_and_default(config,'analysis','combine_controls','False') == 'True')

# output to file vars

CV_WARNING = float(check_and_default(config,"output","cv_warning","10"))
CV_ERROR = float(check_and_default(config,"output","cv_error","25"))

ZSC_WARNING = float(check_and_default(config,"output","zsc_warning","1.5"))
ZSC_ERROR = float(check_and_default(config,"output","zsc_error","2"))

WARNING_COLOR = openpyxl.styles.Color(rgb=check_and_default(config,"output","WARNING_COLOR",'FFFFA500'))
ERROR_COLOR = openpyxl.styles.Color(rgb=check_and_default(config,"output","ERROR_COLOR",'FFFF0000'))

# debugging vars

VERBOSE_OUTPUT = (check_and_default(config,'debugging','verbose_output','False') == 'True')
INCLUDE_PERPLATE_CONTROLS = (check_and_default(config,'debugging','include_perplate_controls','False') == 'True')
TALLY_PERPLATE_NA = (check_and_default(config,'debugging','tally_perplate_na','False') == 'True')
BEADCOUNT_SHEET = (check_and_default(config,'debugging','beadcount_sheet','False') == 'True')

os.chdir("Luminex Documents")

data = dict()
beadcounts = dict()
beadnames = list()
for num in range(PLATE_COUNT):
    for typ in PLATE_TYPES:
        platename = PROTOCOL_NAME+"_"+str(num+1)+typ+".xlsx"
        print("Adding "+platename)
        wb = openpyxl.load_workbook(platename)
        ws = wb['FI']

        # find correct starting indicies for both FI and beadcount
        currow = 1
        while ws.cell(row = currow, column = 3).value != "Description":
            currow = currow + 1
        currow = currow + 1
        if VERBOSE_OUTPUT:
            print("First row of data is "+str(currow))
        if num == 0:
            for col in ws.iter_cols(min_row=currow-2, max_row=currow-2, min_col=4):
                beadnames.extend([cell.value for cell in col if cell != None])
            if VERBOSE_OUTPUT:
                print("Current bead ID count is "+str(len(beadnames)))
        
        beadrow = 1
        while wb["Bead Count"].cell(row = beadrow, column = 3).value != "Description":
            beadrow = beadrow + 1
        beadrow = beadrow + 1
        while wb["Bead Count"].cell(row = beadrow, column = 3).value != "Description":
            beadrow = beadrow + 1
        beadrow = beadrow + 1
        if VERBOSE_OUTPUT:
            print("First beadcount row is "+str(beadrow))

        # go through and add samples to data()
        perplate_na = 0
        while ws.cell(row = currow, column = 3).value != None:
            clean_name = ws.cell(row = currow, column = 3).value.replace("_","-")
            ID = str(num+1) + "_" + str(clean_name)
            if ID not in data.keys():
                data[ID] = list()
                beadcounts[ID] = list()
            curcol = 4
            while ws.cell(row = currow, column = curcol).value != None:
                beadcounts[ID].append(wb["Bead Count"].cell(row = beadrow, column = curcol).value)
                if wb["Bead Count"].cell(row = beadrow, column = curcol).value is not None and wb["Bead Count"].cell(row = beadrow, column = curcol).value < BEAD_CUTOFF:
                    if "Control" in str(ws.cell(row = currow, column = 3).value):
                        if wb["Bead Count"].cell(row = beadrow+1, column = curcol).value < BEAD_CUTOFF:
                            data[ID].append("NA")
                            perplate_na = perplate_na + 1
                            if VERBOSE_OUTPUT:
                                print(ID+" column number "+str(curcol)+" below bead threshold of "+str(BEAD_CUTOFF))
                        else:
                            data[ID].append(ws.cell(row = currow, column = curcol).value)
                    else:
                        data[ID].append("NA")
                        perplate_na = perplate_na + 1
                        if VERBOSE_OUTPUT:
                            print(ID+" column number "+str(curcol)+" below bead threshold of "+str(BEAD_CUTOFF))
                        
                else:
                    if is_number(ws.cell(row = currow, column = curcol).value):
                        data[ID].append(ws.cell(row = currow, column = curcol).value)
                    else:
                        if VERBOSE_OUTPUT:
                            print(ID+" column number "+str(curcol)+" is NaN")
                        data[ID].append("NA")
                    
                curcol = curcol + 1
            beadrow = beadrow + len(ws.cell(row = currow, column =2).value.split(","))
            currow = currow + 1
            # need to add rows equal to the number of wells being averaged, otherwise procession will be uneven
        if(TALLY_PERPLATE_NA):
            print("Plate "+platename+" low beadcounts: "+str(perplate_na))

os.chdir("..")
print("Generating combined workbook")

wb = openpyxl.Workbook()
ws = wb.worksheets[0]
ws.title = "Combined"

for i in range(len(beadnames)):
    ws.cell(column = i+2, row = 1).value = beadnames[i]
    ws.cell(column = i+2, row = 1).font = openpyxl.styles.Font(bold=True)

currow = 2
for key in data.keys():
    ws.cell(column = 1, row = currow).value = key
    for i in range(len(data[key])):
        ws.cell(column = 2+i, row = currow).value = data[key][i]
    currow = currow + 1

if(BEADCOUNT_SHEET):
    ws = wb.create_sheet("beadcounts")

    for i in range(len(beadnames)):
        ws.cell(column = i+2, row = 1).value = beadnames[i]
        ws.cell(column = i+2, row = 1).font = openpyxl.styles.Font(bold=True)

    currow = 2
    for key in data.keys():
        ws.cell(column = 1, row = currow).value = key
        for i in range(len(data[key])):
            ws.cell(column = 2+i, row = currow).value = beadcounts[key][i]
            if(not is_number(ws.cell(column = 2+i, row = currow).value) or ws.cell(column = 2+i, row = currow).value < BEAD_CUTOFF):
                ws.cell(column = 2+i, row = currow).font = openpyxl.styles.Font(color=ERROR_COLOR)
        currow = currow + 1
    

print("Generating CVs sheet")
ws = wb.create_sheet("CVs")

print("Compiling control data")

controls = dict()
for key in data.keys():
    if CONTROL_NAMES is False:
        if "Control" in key or "control" in key:
            plate = int(key.split("_")[0])
            ID = key.split("_")[1]
            if COMBINE_CONTROLS is True:
                ID = ID.split("-")[0]
            if ID not in controls.keys():
                if COMBINE_CONTROLS is True:
                    controls[ID] = list()
                else:
                    controls[ID] = [[] for i in range(PLATE_COUNT)]
            if COMBINE_CONTROLS is True:
                controls[ID].append(data[key])
            else:
                controls[ID][plate-1]=data[key]
    else:
        if len([i for i in CONTROL_NAMES if i in key]) > 0:
            plate = int(key.split("_")[0])
            ID = key.split("_")[1]
            if ID not in controls.keys():
                controls[ID] = [[] for i in range(PLATE_COUNT)]
            controls[ID][plate-1]=data[key]

if VERBOSE_OUTPUT:
    print(str(len(controls.keys()))+" controls found")

for i in range(len(beadnames)):
    ws.cell(column = i+2, row = 1).value = beadnames[i]
    ws.cell(column = i+2, row = 1).font = openpyxl.styles.Font(bold=True)
if(INCLUDE_PERPLATE_CONTROLS):
    ws.cell(column = len(beadnames)+2, row = 1).value = "ERROR COUNT"
    ws.cell(column = len(beadnames)+2, row = 1).font = openpyxl.styles.Font(bold=True)
    ws.cell(column = len(beadnames)+3, row = 1).value = "ERROR+WARNING COUNT"
    ws.cell(column = len(beadnames)+3, row = 1).font = openpyxl.styles.Font(bold=True)

currow = 2
for key in controls.keys():
    per_bead = transpose(controls[key])
    temp_bead = list()
    for i in range(len(beadnames)):
        temp_bead.append([float(z) for z in per_bead[i] if z != "NA"])
    if(INCLUDE_PERPLATE_CONTROLS):
        for plate in range(len(per_bead[0])):
            zsc_err_count = 0
            zsc_war_count = 0
            ws.cell(column = 1, row = currow).value = "Plate " + str(plate+1)
            for bead in range(len(per_bead)):
                ws.cell(column = bead + 2, row = currow).value = per_bead[bead][plate]
                if(per_bead[bead][plate]!='NA'):
                    zscore = abs(float(per_bead[bead][plate])-average(temp_bead[bead]))/std(temp_bead[bead])
                    if zscore > ZSC_ERROR:
                        zsc_err_count = zsc_err_count + 1
                        ws.cell(column = bead + 2, row = currow).font = openpyxl.styles.Font(color=ERROR_COLOR)
                    elif zscore > ZSC_WARNING:
                        zsc_war_count = zsc_war_count + 1
                        ws.cell(column = bead + 2, row = currow).font = openpyxl.styles.Font(color=WARNING_COLOR)
                else:
                    ws.cell(column = bead + 2, row = currow).font = openpyxl.styles.Font(color=ERROR_COLOR)
            ws.cell(column = len(beadnames)+2, row = currow).value = zsc_err_count
            ws.cell(column = len(beadnames)+3, row = currow).value = zsc_war_count + zsc_err_count
            currow = currow + 1
    ws.cell(column = 1, row = currow).value = key
    ws.cell(column = 1, row = currow).font = openpyxl.styles.Font(bold=True)
    for i in range(len(beadnames)):
        CV = std(temp_bead[i])/average(temp_bead[i])*100
        ws.cell(column = i+2, row = currow).value = CV
        if CV >= CV_ERROR:
            ws.cell(column = i+2, row = currow).font = openpyxl.styles.Font(color=ERROR_COLOR, bold=INCLUDE_PERPLATE_CONTROLS)
        elif CV >= CV_WARNING:
            ws.cell(column = i+2, row = currow).font = openpyxl.styles.Font(color=WARNING_COLOR, bold=INCLUDE_PERPLATE_CONTROLS)
        else:
            ws.cell(column = i+2, row = currow).font = openpyxl.styles.Font(bold=INCLUDE_PERPLATE_CONTROLS)
    currow = currow + 1

print("Generating master sheet")
ws = wb.create_sheet("master")

numbers = list()
samples = list()
for key in data.keys():
    if not ((CONTROL_NAMES is False and ("Control" in key or "control" in key)) or (CONTROL_NAMES is not False and (len([i for i in CONTROL_NAMES if i in key]) > 0))):
        samples.append(key.split("_")[1])
        numbers.append(data[key])

for i in range(len(samples)):
    ws.cell(column = 1, row = i+2).value = samples[i]
    ws.cell(column = 1, row = i+2).font = openpyxl.styles.Font(bold=True)

for i in range(len(beadnames)):
    ws.cell(column = i+2, row = 1).value = beadnames[i]
    ws.cell(column = i+2, row = 1).font = openpyxl.styles.Font(bold=True)

for col in range(len(numbers)):
    for row in range(len(numbers[col])):
        ws.cell(column = row+2, row = col+2).value = numbers[col][row]

wb.save(PROTOCOL_NAME+"_results_combined.xlsx")
print(PROTOCOL_NAME+"_results_combined.xlsx saved")

while(len(wb.worksheets) > 1):
    wb.remove_sheet(wb.worksheets[0])
wb.save(PROTOCOL_NAME+"_results_master.xlsx")
print(PROTOCOL_NAME+"_results_master.xlsx saved")
